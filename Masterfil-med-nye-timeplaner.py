from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import csv
from datetime import datetime, date, timedelta

#Navnet på filen som lastes inn
wb = load_workbook('Timesheets.xlsx')
ws = wb.active



last_day_of_prev_month = date.today().replace(day=1) - timedelta(days=1)
start_day_of_prev_month = date.today().replace(day=1) - timedelta(days=last_day_of_prev_month.day)
tenth_day_prev_month = start_day_of_prev_month + timedelta(days=9)

last_day_of_prev_month = datetime.strftime(last_day_of_prev_month, '%Y%m%d')
start_day_of_prev_month = datetime.strftime(start_day_of_prev_month, '%Y%m%d')
tenth_day_prev_month = datetime.strftime(tenth_day_prev_month, '%Y%m%d')


for index, row in enumerate(ws.rows, start=1):
    ws.cell(row=index, column=18).value = ''


    #Følgende går igjennom Trenere og Ungdomsgrupper
    if ws.cell(row=index, column=13).value in ['Trenere - barne og ungdomsgrupper']:
        if ws.cell(row=index, column=3).value:
            ws.cell(row=index, column=1).value = ws.cell(row=index, column=3).value
        else:
            ws.cell(row=index, column=1).value = ws.cell(row=index, column=1).value +' '+ ws.cell(row=index, column=2).value
            print(ws.cell(row=index, column=1).value)

        cur_val = ws.cell(row=index, column=1).value # setting to make if-statement shorter
        if ws.cell(row=index, column=9).value == 280:#Trener 2 !!Merk hvis verdi starter med 0 (eks: 002) er det string, starter det med 1 (eks 101) må det sjekkes som int
            ws.cell(row=index, column=2).value = '0108-F' #Trener 2 m/FP
        elif ws.cell(row=index, column=9).value == 240: #Trener 1
            ws.cell(row=index, column=2).value = '0107-F' #Trener 1 m/FP
        elif ws.cell(row=index, column=9).value == 185: #KLI under 18
            ws.cell(row=index, column=2).value = '0113-F' #under 18
        else: #Klatreleder inne
            ws.cell(row=index, column=2).value = '0111-F' #KLI m/FP

        ws.cell(row=index, column=17).value = ws.cell(row=index, column=8).value
        ws.cell(row=index, column=5).value = 'Monthly'
        ws.cell(row=index, column=6).value = 'Monthly'
        ws.cell(row=index, column=7).value = tenth_day_prev_month
        ws.cell(row=index, column=8).value = start_day_of_prev_month
        ws.cell(row=index, column=9).value = last_day_of_prev_month
        ws.cell(row=index, column=13).value = 3 # 3 for VV, 2 for brygg, 1 for bh
        ws.cell(row=index, column=14).value = '9045' #spørre/sjekke mail
        ws.cell(row=index, column=24).value = 'x'


    #Følgende elif går igjennom kurs
    elif ws.cell(row=index, column=13).value in ['Kurs - Timeplan']:
        if ws.cell(row=index, column=3).value:
            ws.cell(row=index, column=1).value = ws.cell(row=index, column=3).value
        else:
            ws.cell(row=index, column=1).value = ws.cell(row=index, column=1).value +' '+ ws.cell(row=index, column=2).value
            print(ws.cell(row=index, column=1).value)

        if ws.cell(row=index, column=9).value == 240: #or ws.cell(row=index, column=2).value == 056 or ws.cell(row=index, column=2).value == 056
            ws.cell(row=index, column=2).value = '0107-F' #trener 1, samme som sport 1??
        elif ws.cell(row=index, column=9).value == 280:
            ws.cell(row=index, column=2).value = '0108-F' #Sport 2 eller Trener 2
        elif ws.cell(row=index, column=9).value == 185:
            ws.cell(row=index, column=2).value = '0113-F' #KLI under 18
        else:
            ws.cell(row=index, column=2).value = '0102-F' # KLI: 02 F Normaltid Høy Sats m/FP

        ws.cell(row=index, column=17).value = ws.cell(row=index, column=8).value
        ws.cell(row=index, column=5).value = 'Monthly'
        ws.cell(row=index, column=6).value = 'Monthly'
        ws.cell(row=index, column=7).value = tenth_day_prev_month
        ws.cell(row=index, column=8).value = start_day_of_prev_month
        ws.cell(row=index, column=9).value = last_day_of_prev_month
        ws.cell(row=index, column=13).value = 3 # 3 for VV, 2 for brygg, 1 for bh
        ws.cell(row=index, column=14).value = 9003 #prosjektnummer for innføringskurs
        ws.cell(row=index, column=24).value = 'x'


    #Følgende elif går igjennom resepsjonsvakttimeplanen
    elif ws.cell(row=index, column=13).value in ['Resepsjonsvakt - Timeplan']:
        if ws.cell(row=index, column=3).value:
            ws.cell(row=index, column=1).value = ws.cell(row=index, column=3).value
        else:
            ws.cell(row=index, column=1).value = ws.cell(row=index, column=1).value +' '+ ws.cell(row=index, column=2).value
            print(ws.cell(row=index, column=1).value)

        if ws.cell(row=index, column=9).value == 185: #or ws.cell(row=index, column=2).value == 056 or ws.cell(row=index, column=2).value == 056
            ws.cell(row=index, column=2).value = '0113-F' #under 18
        else:
            ws.cell(row=index, column=2).value = '0112-F'

        ws.cell(row=index, column=17).value = ws.cell(row=index, column=8).value
        ws.cell(row=index, column=5).value = 'Monthly'
        ws.cell(row=index, column=6).value = 'Monthly'
        ws.cell(row=index, column=7).value = tenth_day_prev_month
        ws.cell(row=index, column=8).value = start_day_of_prev_month
        ws.cell(row=index, column=9).value = last_day_of_prev_month
        ws.cell(row=index, column=13).value = 3 # 3 for VV, 2 for brygg, 1 for Bergenshallen
        ws.cell(row=index, column=14).value = '9045' #spørre/sjekke mail
        ws.cell(row=index, column=24).value = 'x'


    #Følgende elif tar for seg rutesetting timeplan
    elif ws.cell(row=index, column=13).value in ['Rutesetting - Timeplan']:
        if ws.cell(row=index, column=3).value: #Sjekker om det er et ansattnummer
            ws.cell(row=index, column=1).value = ws.cell(row=index, column=3).value
        else:       #printer navnet hcis de ikke har et ansattnummer registrert
            ws.cell(row=index, column=1).value = ws.cell(row=index, column=1).value +' '+ ws.cell(row=index, column=2).value
            print(ws.cell(row=index, column=1).value)

        if ws.cell(row=index, column=9).value == 215:
            ws.cell(row=index, column=2).value = '0101-F'   #Normaltid ordinær m/FP
        elif ws.cell(row=index, column=9).value == 200:
            ws.cell(row=index, column=2).value = '0112-F'
        else:
            print(f'Rutesetter med annen lønn på linje {index}')

        ws.cell(row=index, column=17).value = ws.cell(row=index, column=8).value
        ws.cell(row=index, column=5).value = 'Monthly'
        ws.cell(row=index, column=6).value = 'Monthly'
        ws.cell(row=index, column=7).value = tenth_day_prev_month
        ws.cell(row=index, column=8).value = start_day_of_prev_month
        ws.cell(row=index, column=9).value = last_day_of_prev_month
        ws.cell(row=index, column=13).value = 3 # 3 for VV, 2 for brygg, 1 for bh
        ws.cell(row=index, column=14).value = '9045' #spørre/sjekke mail
        ws.cell(row=index, column=24).value = 'x'


    #Følgende elif tar for seg AKTIVITET PÅ DAGTID timeplan
    elif ws.cell(row=index, column=13).value in ['Aktivitet på dagtid']:
        if ws.cell(row=index, column=3).value: #Sjekker om det er et ansattnummer
            ws.cell(row=index, column=1).value = ws.cell(row=index, column=3).value
        else:       #printer navnet hcis de ikke har et ansattnummer registrert
            ws.cell(row=index, column=1).value = ws.cell(row=index, column=1).value +' '+ ws.cell(row=index, column=2).value
            print(ws.cell(row=index, column=1).value)

        if ws.cell(row=index, column=9).value == 240: #or ws.cell(row=index, column=2).value == 056 or ws.cell(row=index, column=2).value == 056
            ws.cell(row=index, column=2).value = '0107-F' #trener 1, samme som sport 1??
        elif ws.cell(row=index, column=9).value == 280:
            ws.cell(row=index, column=2).value = '0108-F' #Sport 2 eller Trener 2
        elif ws.cell(row=index, column=9).value == 185:
            ws.cell(row=index, column=2).value = '0113-F' #KLI under 18
        else:
            ws.cell(row=index, column=2).value = '0102-F' # KLI: 02 F Normaltid Høy Sats m/FP

        ws.cell(row=index, column=17).value = ws.cell(row=index, column=8).value
        ws.cell(row=index, column=5).value = 'Monthly'
        ws.cell(row=index, column=6).value = 'Monthly'
        ws.cell(row=index, column=7).value = tenth_day_prev_month
        ws.cell(row=index, column=8).value = start_day_of_prev_month
        ws.cell(row=index, column=9).value = last_day_of_prev_month
        ws.cell(row=index, column=13).value = 3 # 3 for VV, 2 for brygg, 1 for bh
        ws.cell(row=index, column=14).value = '9045' #spørre/sjekke mail
        ws.cell(row=index, column=24).value = 'x'


    #Følgende elif tar for seg VASK I br timeplan
    elif ws.cell(row=index, column=13).value in ['Vask Bryggeriet - Timeplan']:
        if ws.cell(row=index, column=3).value: #Sjekker om det er et ansattnummer
            ws.cell(row=index, column=1).value = ws.cell(row=index, column=3).value
        else:       #printer navnet hcis de ikke har et ansattnummer registrert
            ws.cell(row=index, column=1).value = ws.cell(row=index, column=1).value +' '+ ws.cell(row=index, column=2).value
            print(ws.cell(row=index, column=1).value)

        if ws.cell(row=index, column=9).value == 200:
            ws.cell(row=index, column=2).value = '0103-F'   #Normaltid ordinær m/FP
        else:
            print(f'Vask i Bryggeriet med annen lønn på linje {index}')

        ws.cell(row=index, column=17).value = ws.cell(row=index, column=8).value
        ws.cell(row=index, column=5).value = 'Monthly'
        ws.cell(row=index, column=6).value = 'Monthly'
        ws.cell(row=index, column=7).value = tenth_day_prev_month
        ws.cell(row=index, column=8).value = start_day_of_prev_month
        ws.cell(row=index, column=9).value = last_day_of_prev_month
        ws.cell(row=index, column=13).value = 3 # 3 for VV, 2 for brygg, 1 for bh
        ws.cell(row=index, column=14).value = '9045' #spørre/sjekke mail
        ws.cell(row=index, column=24).value = 'x'


    #i tilfelle det ikke høre til en timeplan
    else:
        print('Timeplan??')
        print(f'feil på rad nr {index}')

# Sletter kolonner som ikke trengs, og legger til tomme kolonner
ws.delete_cols(3,2)
ws.insert_cols(3,2)
ws.delete_cols(12)
ws.insert_cols(12)
ws.delete_cols(15,2)
ws.insert_cols(15,2)




#Endrer overskriften på kolonnene til å stemme overens med rett format
ws.delete_rows(1)
ws.insert_rows(1)
ws['A1']='Employee'
ws['B1']='PD'
ws['C1']='Text'
ws['D1']='Internal info'
ws['E1']='PayrollRun'
ws['F1']='Frequency setup'
ws['G1']='Next Date'
ws['H1']='Date From'
ws['I1']='Date To'
ws['J1']='Supplier'
ws['K1']='InvoiceNo'
ws['L1']='Xidentifier'
ws['M1']='Object 1'
ws['N1']='Object 2'
ws['O1']='XGL'
ws['P1']='ObjectValue'
ws['Q1']='Rate 1'
ws['R1']='Rate 2'
ws['S1']='Rate 3'
ws['T1']='Rate 4'
ws['U1']='Rate 5'
ws['V1']='Rate 6'
ws['W1']='Rate 7'
ws['X1']='x'

#Lagrer den aktive sheeten som csv
sh = wb.active # was .get_active_sheet()
d_tday = datetime.today()
with open(f"PR02_Timelister - Kurs Resepsjon Treningsgrupper - {d_tday.strftime('%Y %m')}.csv", 'w', newline="") as file_handle:
    csv_writer = csv.writer(file_handle, delimiter = ";")
    for row in sh.iter_rows(): # generator; was sh.rows
        csv_writer.writerow([cell.value for cell in row])

#Lagrer endringene til en ny excelfil med navnet gitt under
wb.save(f"Timelister - Kurs Resepsjon Treningsgrupper - {d_tday.strftime('%Y %m')}.xlsx")
